"""Excel 文件读写模块。"""

import os
import copy
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def read_excel(path):
    """读取 Excel 文件，提取前两列（中文原文、英文译文）。

    Args:
        path: Excel 文件路径

    Returns:
        list[dict]: [{"row": 1, "source": "中文", "target": "English"}, ...]
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    data = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过空行
        if not row or len(row) < 2:
            continue
        source = str(row[0]).strip() if row[0] is not None else ""
        target = str(row[1]).strip() if row[1] is not None else ""
        if source or target:
            data.append({"row": idx, "source": source, "target": target})

    wb.close()
    logger.info(f"读取了 {len(data)} 行数据 (从 {path})")
    return data


def write_results_to_excel(original_path, results, output_path):
    """在原始 Excel 基础上追加结果列，保存为新文件。

    Args:
        original_path: 原始 Excel 文件路径
        results: 结果列表 [{"row": 2, "source": ..., "target": ..., "result": {...}}, ...]
        output_path: 输出文件路径
    """
    wb = load_workbook(original_path)
    ws = wb.active

    # 找到最后一列
    max_col = ws.max_column

    # 添加表头
    headers = ["评分", "问题", "修改建议", "总结"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for i, header in enumerate(headers):
        cell = ws.cell(row=1, column=max_col + 1 + i, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 构建行号到结果的映射
    result_map = {}
    for r in results:
        result_map[r["row"]] = r.get("result", {})

    # 写入结果数据
    for row_num, result in result_map.items():
        score = result.get("score", "")
        issues = "\n".join(result.get("issues", []))
        suggestion = result.get("suggestion", "")
        summary = result.get("summary", "")

        values = [score, issues, suggestion, summary]
        for i, val in enumerate(values):
            cell = ws.cell(row=row_num, column=max_col + 1 + i, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

            # 低分标红
            if i == 0 and isinstance(score, int) and score <= 5:
                cell.font = Font(color="FF0000", bold=True)

    # 调整列宽
    for i in range(len(headers)):
        col_letter = get_column_letter(max_col + 1 + i)
        ws.column_dimensions[col_letter].width = 30 if i > 0 else 8

    wb.save(output_path)
    wb.close()
    logger.info(f"结果已写入: {output_path}")


def write_independent_report(data, results, output_path):
    """生成独立的结果报告 Excel。

    Args:
        data: 原始数据列表
        results: 结果列表
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "翻译校验报告"

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    headers = ["行号", "中文原文", "英文译文", "评分", "问题", "修改建议", "总结"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 构建结果映射
    result_map = {}
    for r in results:
        result_map[r["row"]] = r.get("result", {})

    # 写入数据
    for row_idx, item in enumerate(data, start=2):
        result = result_map.get(item["row"], {})
        score = result.get("score", "")
        issues = "\n".join(result.get("issues", []))
        suggestion = result.get("suggestion", "")
        summary = result.get("summary", "")

        row_data = [item["row"], item["source"], item["target"],
                     score, issues, suggestion, summary]

        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

        # 根据评分设置行背景色
        if isinstance(score, int):
            if score <= 5:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = red_fill
            elif score <= 7:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = yellow_fill

    # 设置列宽
    widths = [6, 40, 40, 8, 30, 30, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(output_path)
    wb.close()
    logger.info(f"独立报告已生成: {output_path}")
